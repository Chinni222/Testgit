import boto3
import openpyxl
import os
import datetime

def get_ec2_instances(region_name, access_key, secret_key):
    try:
        session = boto3.Session(
            aws_access_key_id=access_key,
            aws_secret_access_key=secret_key,
            region_name=region_name
        )

        ec2 = session.resource('ec2')
        instances = ec2.instances.filter(Filters=[{'Name': 'instance-state-name', 'Values': ['running', 'stopped']}])

        instance_data = []
        for instance in instances:
            tags = instance.tags or []
            name_tag = next((tag for tag in tags if tag['Key'] == 'Name'), None)
            name = name_tag['Value'] if name_tag else "Unknown"

            # Get platform and architecture details
            platform = instance.platform
            architecture = instance.architecture

            instance_info = {
                'Name': name,
                'Instance ID': instance.id,
                'Instance Type': instance.instance_type,
                'State': instance.state['Name'],
                'Public IP': instance.public_ip_address,
                'Private IP': instance.private_ip_address,
                'Key Name': instance.key_name,
                'Platform': platform,
                'Architecture': architecture
            }
            instance_data.append(instance_info)
        return instance_data

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return []

def create_excel_file(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'EC2 Instances'

    headers = ['Name', 'Instance ID', 'Instance Type', 'State', 'Public IP', 'Private IP', 'Key Name', 'Platform', 'Architecture']

    # Write headers as cell values
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx).value = header

    # Add a blank row
    sheet.cell(row=2, column=1).value = ''  # Empty cell to represent the blank row

    for row_idx, row in enumerate(data, start=3):
        # Write data rows starting from row 3 (after headers and blank row)
        for col_idx, value in enumerate(row.values(), start=1):
            sheet.cell(row=row_idx, column=col_idx).value = value

    # Validate and create the desired location directory if it doesn't exist
    if not os.path.exists('C:/IDI'):
        os.makedirs('C:/IDI')

    # Construct the full filepath
    current_month = datetime.datetime.now().strftime('%B')
    currentMonthFileName = f'{current_month}_INVENTORY_DATA_{datetime.datetime.now().year}.xlsx'
    filepath = os.path.join('C:/IDI', currentMonthFileName)
    workbook.save(filepath)
    print('Inventory Data Created Successfully in:', filepath)

if __name__ == '__main__':
    region_name = input("Enter your Region Name : ")
    access_key = input("Enter your Access Key : ")
    secret_key = input("Enter your Secret Key : ")

    instance_data = get_ec2_instances(region_name, access_key, secret_key)
    create_excel_file(instance_data)